import re
import ccxt
import os
from telethon.sync import TelegramClient
from telethon import events
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime
import array as arr
import BinanceBot
import WhaleHunterFormat    
# ---------------------------------------------INFO_TOKEN_GET_API-------------------------------------------------------------
api_id = 27469665
api_hash = 'b1778da8e2e6df457ed506e6a3f0bbf9'
# group_name = 'cointrendz_whalehunter'
group_name = 'linkau01'
# --------------------------------------------CHECK_PREMIUM_GROUP-------------------------------------------------------------
# CHECK FORMAT TOKEN


def checkNotification(message_text):
    if (message_text.upper().find('BUY') != -1 or message_text.upper().find('SELL') != -1):
        if (message_text.upper().find('STOPLOSS') != -1):
            return True
        else:
            return False
    else:
        return False
# END CHECK

def checkVolume(message_text): 
    match = re.search(r"24h Vol: (\d+(\.\d+)?)([KMBT]?)", message_text)
    if match:
        vol_value = match.group(1)
        vol_suffix = match.group(3)
        vol_values = float(vol_value)
        if(vol_values>=600 and vol_suffix.upper()=='K'):
            return True
        if(vol_suffix.upper()=='M'):
            return True

# GET TOKEN
def symbolToken(message_text):
    tokenCoin = ''
    checkTokenFormat = False
    Token = ''
    for char in message_text:
        if (char == '\n'):
            break
        else:
            tokenCoin += char
        if (ord(char) == 35):
            checkTokenFormat = True
    if (checkTokenFormat == True and tokenCoin.upper().find('USDT') != -1 and checkNotification(message_text)):
        Token = re.search(r"\b([A-Z]+)\b", tokenCoin).group()
        return Token+'/USDT'
    return False
# END GET TOKEN


# ------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------CREATE_DATABASE----------------------------------------------------------
def createDatabaseToExcel(token, entry, sell, message):
    timeNow = datetime.now()
    file_path = "E:/Python/database.xlsx"
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        sheet = wb.active
        data = [
            [token, entry, sell, timeNow, message]]
    else:
        wb = Workbook()
        sheet = wb.active
        data = [
            ["TOKEN NAME", "ENTRY BUY", "SELL", "DATE", "MESSAGE"],
            [token, entry, sell, timeNow, message]]
    alignment = Alignment(horizontal="center",
                          vertical="center", wrap_text=True)
    for row in data:
        sheet.append(row)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment
    wb.save("database.xlsx")
# ------------------------------------------------------------------------------------------------------------------------------
# ------------------------------------------GET_PRICE_BUY_SELL----------------------------------------------------------------------------


def findPrice_Buy(message_text):
   # Tìm vị trí của chuỗi "Buy :"
    buy_index = message_text.find("Buy :")
    if buy_index != -1:
        # Tìm vị trí của ký tự ":" sau chuỗi "Buy :"
        colon_index = message_text.find(":", buy_index)
     # Tìm vị trí của ký tự "\n" sau chuỗi "Buy :"
        newline_index = message_text.find("\n", colon_index)
    # Tách chuỗi "buy" từ vị trí sau ":" đến trước "\n"
        buy = (message_text[colon_index + 1:newline_index].strip())
        return buy
    return ''


def findPrice_Sell(message_text):
    # Tìm vị trí của chuỗi "Sell :"
    sell_index = message_text.find("Sell :")
    if sell_index != -1:
        # Tìm vị trí của ký tự ":" sau chuỗi "Sell :"
        colon_index = message_text.find(":", sell_index)
    # Tìm vị trí của ký tự "\n" sau chuỗi "Sell :"
        newline_index = message_text.find("\n", colon_index)
    # Tách chuỗi "sell" từ vị trí sau ":" đến trước "\n"
        sell = str(message_text[colon_index + 1:newline_index].strip())
        return sell
    return ''
# ------------------------------------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------------------------------------

#  ------------------------------------------------------------------------------------------------------------------------------

# MAIN_CHECK


checkCommand = False


async def main():
    async with TelegramClient('session_names', api_id, api_hash) as client:
        @client.on(events.NewMessage(chats=group_name))
        async def handle_new_message(event):
            global checkCommand
            text = event.message.text
            message_text = str(event.message.text)
            if (WhaleHunterFormat.checkSymbolToken(message_text) != False):
                symbolToken_Buy = WhaleHunterFormat.checkSymbolToken(
                    message_text)
                print("---------------------------- LỆNH MỚI - NEW COMMAND ----------------------------------------------------------")
                print("TOKEN BUY: "+str(symbolToken_Buy))
                if(checkVolume(message_text)):
                    BinanceBot.setCommandBuyLimit(symbolToken_Buy)
                else:
                    print("VOLUME KHÔNG ĐẠT !!!")
            # print(message_text)
            # if (symbolToken(message_text) != False):
            #     symbolToken_Buy = symbolToken(message_text)
            #     print(symbolToken_Buy)
            #     if (checkCommand == False):
            #         BinanceBot.setCommandBuyLimit(
            #             symbolToken_Buy, BinanceBot.currentPrice(symbolToken_Buy))
            #         checkCommand = True
            # createDatabaseToExcel(symbolToken(
            #     message_text), findPrice_Buy(message_text), findPrice_Sell(message_text), message_text)
        await client.run_until_disconnected()
# END_CHECK


if __name__ == '__main__':
    import asyncio
    asyncio.run(main())
